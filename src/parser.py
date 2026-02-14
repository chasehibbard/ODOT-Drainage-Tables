"""
parser.py — Reads raw OpenRoads Designer flex table exports (.xlsx)
and returns clean Python data structures for the formatters.
"""

import openpyxl


def _find_column(header, target_phrases):
    """
    Find a column index by checking if any of the target phrases
    appear in the header text. Checks in order — first match wins.
    """
    for i, h in enumerate(header):
        h_upper = h.upper()
        for phrase in target_phrases:
            if phrase in h_upper:
                return i
    return None


def parse_da_summary(filepath):
    """
    Reads the OKDOT_Drainage Areas Summary Table.xlsx export.

    Returns a list of dicts, one per drainage area:
        {
            "designation": "DA-2",
            "area_acres": 0.019,
            "tc_min": 10,
            "runoff_coeff": 0.25,
            "intensity": 9.058,
            "flow_cfs": 0.05,
        }
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()

    if not rows:
        return []

    header = [str(h).strip() if h else "" for h in rows[0]]

    # Map columns using specific phrases (checked in priority order)
    col_map = {
        "designation": _find_column(header, ["DRAINAGE AREA DESIGNATION"]),
        "outflow_structure": _find_column(header, ["OUTFLOW STRUCTURE"]),
        "area_acres": _find_column(header, ["DRAINAGE AREA (ACRE", "DRAINAGE AREA(ACRE"]),
        "tc_min": _find_column(header, ["TIME OF CONCENTRATION"]),
        "runoff_coeff": _find_column(header, ["RUNOFF COEFFICIENT"]),
        "intensity": _find_column(header, ["CATCHMENT INTENSITY", "INTENSITY"]),
        "flow_cfs": _find_column(header, ["CATCHMENT RATIONAL FLOW", "RATIONAL FLOW"]),
    }

    # Remove any keys that didn't find a match
    col_map = {k: v for k, v in col_map.items() if v is not None}

    data = []
    for row in rows[1:]:
        if row[0] is None:
            continue

        record = {}
        for key, col_idx in col_map.items():
            val = row[col_idx] if col_idx < len(row) else None
            record[key] = val

        data.append(record)

    return data


def parse_inlets(filepath):
    """
    Reads the OKDOT_Inlets for Storm Sewer Design Record.xlsx export.

    Returns a list of dicts, one per inlet structure.
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()

    if not rows:
        return []

    header = [str(h).strip() if h else "" for h in rows[0]]

    col_map = {
        "structure_no": _find_column(header, ["STRUCTURE NO"]),
        "alignment": _find_column(header, ["ALIGNMENT"]),
        "station": _find_column(header, ["STATION"]),
        "offset": _find_column(header, ["OFFSET"]),
        "structure_type": _find_column(header, ["STRUCTURE TYPE"]),
        "description": _find_column(header, ["DESCRIPTION"]),
        "elevation_rim": _find_column(header, ["ELEVATION (RIM)", "ELEVATION"]),
        "flow_captured": _find_column(header, ["FLOW (CAPTURED)", "CAPTURED"]),
        "flow_bypassed_rational": _find_column(header, ["BYPASSED RATIONAL"]),
        "flow_total_bypassed": _find_column(header, ["TOTAL BYPASSED"]),
        "capture_efficiency": _find_column(header, ["CAPTURE EFFICIENCY"]),
        "max_spread": _find_column(header, ["SPREAD"]),
        "hgl_out": _find_column(header, ["HYDRAULIC GRADE"]),
        "egl_out": _find_column(header, ["ENERGY GRADE"]),
        "bypass_target": _find_column(header, ["BYPASS TARGET"]),
    }

    col_map = {k: v for k, v in col_map.items() if v is not None}

    data = []
    for row in rows[1:]:
        if row[0] is None:
            continue

        record = {}
        for key, col_idx in col_map.items():
            val = row[col_idx] if col_idx < len(row) else None
            record[key] = val

        data.append(record)

    return data
