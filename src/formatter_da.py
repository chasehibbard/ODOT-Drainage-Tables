"""
formatter_da.py — Builds the formatted ODOT Drainage Area Summary Table.

Replicates the layout and formatting from the ODOT output example:
  - Row 1: blank
  - Row 2: "DRAINAGE AREA SUMMARY" title merged across A-L
  - Row 3: Primary headers (with merged cells for grouped columns)
  - Row 4: Sub-headers (10/25/50/100 YEAR under intensity and flow)
  - Rows 5+: Data
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Thin border used on all cells
THIN_SIDE = Side(style="thin")
ALL_BORDERS = Border(
    top=THIN_SIDE, bottom=THIN_SIDE, left=THIN_SIDE, right=THIN_SIDE
)

HEADER_FONT = Font(name="Arial", size=10)
DATA_FONT = Font(name="Arial", size=10)

CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center")

# Column widths matching the ODOT example
COL_WIDTHS = {
    "A": 15.55, "B": 11.55, "C": 10.33, "D": 10.89,
    "E": 13, "F": 13, "G": 13, "H": 13,
    "I": 13, "J": 13, "K": 13, "L": 13,
}

# Row heights matching the ODOT example
ROW_HEIGHTS = {
    2: 15.75,  # title
    3: 18.75,  # primary headers
    4: 24.75,  # sub-headers
}
DATA_ROW_HEIGHT = 15.75


def create_da_summary_sheet(wb, da_data):
    """
    Creates the Drainage Area Summary sheet in the given workbook.

    Args:
        wb: openpyxl Workbook
        da_data: list of dicts from parser.parse_da_summary()

    Returns:
        The created worksheet
    """
    ws = wb.create_sheet(title="Drainage Area Summary")

    # --- Column widths ---
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # --- Row 1: blank ---

    # --- Row 2: Title ---
    ws.merge_cells("A2:L2")
    title_cell = ws["A2"]
    title_cell.value = "DRAINAGE AREA SUMMARY"
    title_cell.font = HEADER_FONT
    title_cell.alignment = CENTER_WRAP
    title_cell.border = Border(top=THIN_SIDE)
    ws.row_dimensions[2].height = ROW_HEIGHTS[2]

    # --- Row 3: Primary headers ---
    ws.row_dimensions[3].height = ROW_HEIGHTS[3]

    # Single-cell headers that span rows 3-4
    single_headers = [
        ("A", "DRAINAGE AREA DESIGNATION"),
        ("B", "DRAINAGE AREA (ACRE)"),
        ("C", "TOTAL TC (MIN)"),
        ("D", 'RUNOFF COEFF. "C"'),
    ]
    for col_letter, text in single_headers:
        merge_range = f"{col_letter}3:{col_letter}4"
        ws.merge_cells(merge_range)
        cell = ws[f"{col_letter}3"]
        cell.value = text
        cell.font = HEADER_FONT
        cell.alignment = CENTER_WRAP
        cell.border = ALL_BORDERS

    # Grouped header: RAINFALL INTENSITY (IN/HR) spanning E-H
    ws.merge_cells("E3:H3")
    cell = ws["E3"]
    cell.value = "RAINFALL INTENSITY (IN/HR)"
    cell.font = HEADER_FONT
    cell.alignment = CENTER_WRAP
    cell.border = ALL_BORDERS

    # Grouped header: PEAK FLOW (CFS) spanning I-L
    ws.merge_cells("I3:L3")
    cell = ws["I3"]
    cell.value = "PEAK FLOW (CFS)"
    cell.font = HEADER_FONT
    cell.alignment = CENTER_WRAP
    cell.border = ALL_BORDERS

    # --- Row 4: Sub-headers ---
    ws.row_dimensions[4].height = ROW_HEIGHTS[4]
    sub_headers = ["10 YEAR", "25 YEAR", "50 YEAR", "100 YEAR"]

    # Under RAINFALL INTENSITY (cols E-H)
    for i, text in enumerate(sub_headers):
        col_letter = get_column_letter(5 + i)  # E=5, F=6, G=7, H=8
        cell = ws[f"{col_letter}4"]
        cell.value = text
        cell.font = HEADER_FONT
        cell.alignment = CENTER_WRAP
        cell.border = ALL_BORDERS

    # Under PEAK FLOW (cols I-L)
    for i, text in enumerate(sub_headers):
        col_letter = get_column_letter(9 + i)  # I=9, J=10, K=11, L=12
        cell = ws[f"{col_letter}4"]
        cell.value = text
        cell.font = HEADER_FONT
        cell.alignment = CENTER_WRAP
        cell.border = ALL_BORDERS

    # --- Data rows ---
    for row_idx, record in enumerate(da_data, start=5):
        ws.row_dimensions[row_idx].height = DATA_ROW_HEIGHT

        # Column A: Drainage Area Designation
        _write_cell(ws, row_idx, 1, record.get("designation"))
        # Column B: Drainage Area (Acre)
        _write_cell(ws, row_idx, 2, record.get("area_acres"))
        # Column C: Total TC (Min)
        _write_cell(ws, row_idx, 3, record.get("tc_min"))
        # Column D: Runoff Coeff "C"
        _write_cell(ws, row_idx, 4, record.get("runoff_coeff"))

        # Column E: 10-Year Rainfall Intensity
        # The raw export only has one storm frequency. We place it in
        # the 10-Year column. The other columns are left blank for the
        # user to fill in if they have data for other frequencies.
        _write_cell(ws, row_idx, 5, record.get("intensity"))
        # Columns F-H: 25/50/100-Year Intensity (blank)
        for col in range(6, 9):
            _write_cell(ws, row_idx, col, None)

        # Column I: 10-Year Peak Flow
        _write_cell(ws, row_idx, 9, record.get("flow_cfs"))
        # Columns J-L: 25/50/100-Year Peak Flow (blank)
        for col in range(10, 13):
            _write_cell(ws, row_idx, col, None)

    return ws


def _write_cell(ws, row, col, value):
    """Write a value to a cell with standard data formatting."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = DATA_FONT
    cell.alignment = CENTER
    cell.border = ALL_BORDERS
    return cell
