"""
formatter_inlet.py — Builds the formatted ODOT Inlet Design Record sheet.

Replicates the layout from the ODOT output example:
  - Row 1: "DRAINAGE STRUCTURE DESIGN RECORD - INLETS" title merged B1:V1
  - Rows 2-4: Header rows (column headers merged vertically across rows 2-4)
  - Row 5+: Data rows with "PROPOSED" label in column A
"""

from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


THIN_SIDE = Side(style="thin")
ALL_BORDERS = Border(
    top=THIN_SIDE, bottom=THIN_SIDE, left=THIN_SIDE, right=THIN_SIDE
)

HEADER_FONT = Font(name="Arial", size=10)
DATA_FONT = Font(name="Arial", size=10)

CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center")

# Headers with 90-degree rotation
CENTER_ROTATED = Alignment(
    horizontal="center", vertical="center", wrap_text=True, text_rotation=90
)

# Column definitions: (letter, header text, width, rotated?)
COLUMNS = [
    ("B", "STRUCTURE NUMBER", 10.44, True),
    ("C", "ALIGNMENT", 6.0, True),
    ("D", "STATION, OFFSET", 16.11, False),
    ("E", "INLET TYPE", 14.33, False),
    ("F", "DESCRIPTION", 16.55, False),
    ("G", "DESIGN YEAR", 8.33, False),
    ("H", "DRAINAGE AREA", 8.11, True),
    ("I", "AREA (ACRE)", 5.89, True),
    ("J", "C VALUE", 13.0, True),
    ("K", "Tc (MINUTES)", 5.66, True),
    ("L", "INTENSITY (IN/HR)", 6.0, True),
    ("M", " FLOW RATE INITIAL (Qi) (CFS)", 5.66, True),
    ("N", 'GRADE "SL" (ft/ft)', 6.33, True),
    ("O", 'CROSS SLOPE "Sx" (ft/ft)', 5.89, True),
    ("P", "CARRY OVER (CFS)", 6.0, True),
    ("Q", "SUM Q AT INLETS (CFS)", 5.44, True),
    ("R", "DEPTH AT INLET (FT)", 7.33, True),
    ("S", "SPREAD AT INLET (FT)", 7.33, True),
    ("T", "BYPASS(CFS)", 13.0, True),
    ("U", "BYPASS TO INLET", 6.33, True),
    ("V", "COMMENTS", 23.44, False),
]

ROW_HEIGHTS = {
    1: 16.5,    # title row
    2: 80.25,   # header row (tall for rotated text)
    3: 18.75,
    4: 27.0,
}
DATA_ROW_HEIGHT = 15.75


def _extract_inlet_type(description):
    """
    Extract a short inlet type name from the full OpenRoads description path.

    Example input:
        "Node\\StormWaterNode\\1_Proposed\\Inlets\\CI - Curb Inlets\\CI Des 2 STD"
    Example output:
        "CI Des 2 STD"
    """
    if not description:
        return ""
    # Take the last segment of the backslash-separated path
    parts = str(description).split("\\")
    return parts[-1].strip()


def _format_station_offset(station, offset):
    """Combine station and offset into a single string."""
    parts = []
    if station is not None:
        parts.append(str(station))
    if offset is not None and offset != 0:
        parts.append(f"Offset: {offset}")
    return ", ".join(parts) if parts else ""


def create_inlet_sheet(wb, inlet_data, da_data=None):
    """
    Creates the Inlet Design Record sheet in the given workbook.

    Args:
        wb: openpyxl Workbook
        inlet_data: list of dicts from parser.parse_inlets()
        da_data: optional list of dicts from parser.parse_da_summary()
                 Used to fill in drainage area info (area, C, Tc, intensity)

    Returns:
        The created worksheet
    """
    ws = wb.create_sheet(title="Inlets")

    # --- Column widths ---
    ws.column_dimensions["A"].width = 4.44
    for col_letter, _, width, _ in COLUMNS:
        ws.column_dimensions[col_letter].width = width

    # --- Row 1: Title ---
    ws.row_dimensions[1].height = ROW_HEIGHTS[1]
    ws.merge_cells("B1:V1")
    title_cell = ws["B1"]
    title_cell.value = "DRAINAGE STRUCTURE DESIGN RECORD - INLETS"
    title_cell.font = HEADER_FONT
    title_cell.alignment = CENTER_WRAP

    # --- Rows 2-4: Headers ---
    # Each column header is merged across rows 2-4
    for row_num in [2, 3, 4]:
        ws.row_dimensions[row_num].height = ROW_HEIGHTS[row_num]

    for col_letter, header_text, _, rotated in COLUMNS:
        merge_range = f"{col_letter}2:{col_letter}4"
        ws.merge_cells(merge_range)
        cell = ws[f"{col_letter}2"]
        cell.value = header_text
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ROTATED if rotated else CENTER_WRAP
        cell.border = ALL_BORDERS

    # --- Build DA lookup (outflow_structure → DA info) ---
    # Each DA record has an "outflow_structure" field (e.g., "A2", "CI-1")
    # that tells us which inlet structure it drains to.
    da_by_outflow = {}
    if da_data:
        for da in da_data:
            outflow = da.get("outflow_structure", "")
            if outflow:
                da_by_outflow[str(outflow).strip()] = da

    # --- Data rows ---
    data_start = 5
    data_end = data_start + len(inlet_data) - 1

    # Merge column A for "PROPOSED" label
    if inlet_data:
        if len(inlet_data) > 1:
            ws.merge_cells(f"A{data_start}:A{data_end}")
        cell = ws[f"A{data_start}"]
        cell.value = "PROPOSED"
        cell.font = Font(name="Calibri", size=10)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", text_rotation=90
        )
        cell.border = Border(top=THIN_SIDE)

    for row_idx, record in enumerate(inlet_data, start=data_start):
        ws.row_dimensions[row_idx].height = DATA_ROW_HEIGHT

        struct_no = record.get("structure_no", "")

        # Try to find matching DA data for this inlet
        da_record = da_by_outflow.get(struct_no)

        # B: Structure Number
        _write_cell(ws, row_idx, 2, struct_no)
        # C: Alignment
        _write_cell(ws, row_idx, 3, record.get("alignment"))
        # D: Station, Offset
        _write_cell(ws, row_idx, 4, _format_station_offset(
            record.get("station"), record.get("offset")
        ))
        # E: Inlet Type (extracted from description path)
        _write_cell(ws, row_idx, 5, _extract_inlet_type(record.get("description")))
        # F: Description
        _write_cell(ws, row_idx, 6, record.get("structure_type"))
        # G: Design Year (not in raw data — left blank)
        _write_cell(ws, row_idx, 7, None)
        # H: Drainage Area (from DA data if linked)
        _write_cell(ws, row_idx, 8, da_record.get("designation") if da_record else None)
        # I: Area (Acre) (from DA data)
        _write_cell(ws, row_idx, 9, da_record.get("area_acres") if da_record else None)
        # J: C Value (from DA data)
        _write_cell(ws, row_idx, 10, da_record.get("runoff_coeff") if da_record else None)
        # K: Tc (Minutes) (from DA data)
        _write_cell(ws, row_idx, 11, da_record.get("tc_min") if da_record else None)
        # L: Intensity (IN/HR) (from DA data)
        _write_cell(ws, row_idx, 12, da_record.get("intensity") if da_record else None)
        # M: Flow Rate Initial Qi (captured flow from inlet data)
        _write_cell(ws, row_idx, 13, record.get("flow_captured"))
        # N: Grade SL (not in raw data — left blank)
        _write_cell(ws, row_idx, 14, None)
        # O: Cross Slope Sx (not in raw data — left blank)
        _write_cell(ws, row_idx, 15, None)
        # P: Carry Over (total bypassed flow)
        _write_cell(ws, row_idx, 16, record.get("flow_total_bypassed"))
        # Q: Sum Q at Inlets (captured + carry over)
        captured = record.get("flow_captured") or 0
        bypassed = record.get("flow_total_bypassed") or 0
        sum_q = captured + bypassed if (captured or bypassed) else None
        _write_cell(ws, row_idx, 17, sum_q)
        # R: Depth at Inlet (not directly in raw data — left blank)
        _write_cell(ws, row_idx, 18, None)
        # S: Spread at Inlet
        _write_cell(ws, row_idx, 19, record.get("max_spread"))
        # T: Bypass (CFS)
        _write_cell(ws, row_idx, 20, record.get("flow_bypassed_rational"))
        # U: Bypass To Inlet
        _write_cell(ws, row_idx, 21, record.get("bypass_target"))
        # V: Comments (blank)
        _write_cell(ws, row_idx, 22, None)

    return ws



def _write_cell(ws, row, col, value):
    """Write a value to a cell with standard data formatting."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = DATA_FONT
    cell.alignment = CENTER
    cell.border = ALL_BORDERS
    return cell
